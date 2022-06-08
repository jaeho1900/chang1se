"""
P_Grammar.py
"""


# -----------------------------
# String
# -----------------------------

A1 = A2, a = 100, 'I need Tiger.'
print(A1, A2, A3, sep='\n')

# 개수--------------
len(a)
a.count(' ')

# 찾기--------------
a.find('e')   # 맨 왼쪽부터 첫번째 인덱스 위치 반환, 없으면 '-1' 반환
a.rfind('e')  # 맨 오른쪽부터 첫번째 인덱스 위치 반환
a.index('k')  # 맨 왼쪽부터 첫번째 인덱스 위치 반환, 없으면 'ValueError' 발생
a.rindex('e')
[i for i, x in enumerate(a) if x == 'e']  # 모든 원소의 인덱스 위치 반환

# 공백 제거--------------
c = '   white, black, blue   '
c.strip()
c.lstrip()
c.rstrip()

# 좌우 정렬--------------
f = '123'
f.ljust(10, '-')   # 총길이 10, 왼쪽정렬, 나머지 '지정문자' 채움
f.rjust(10, ' ')   # 총길이 10, 오른쪽정렬, 나머지 '공백' 채움

# 알파벳 대문자 바꾸기--------------
'hello, pin-up, world!'.upper()       # 문자열의 모든 알파벳을 대문자로 바꾸기
'hello, pin-up, world!'.title()       # 알파벳 외 모든 문자열의 첫글자를 대문자로 바꾸기
'hello, pin-up, world!'.capitalize()  # 문장의 첫글자만 대문자로 바꾸기

# replace, split, join--------------
a = 'asiana air com'
b = a.replace(' ', '.')
c = b.split('.', maxsplit=1)
d = '@'.join(c)
print(d)

# print 방법 변화--------------
a = 'I need Tiger.'
b = 1233.84592

# str concatenation : + ,
print(a[7:12] + '의 투자비중은' + str(round(b, 2)) + '%입니다.')
print(a[7:12], '의 투자비중은', str(round(b, 2)), '%입니다.', sep='', end='\n\n')

# C 언어 % 타입 : %s, %c, %d, %f, %%
print('value is { %d }' % b)
print('%s의 투자비중은 %-15.2f%%입니다.' % (a[-6:-1], b))        # 좌정렬 및 자릿수 맞춤
print('%s의 투자비중은 % 15.2f%%입니다.' % (a[-6:-1], b))        # 우정렬 및 자릿수 맞춤
print('%s의 투자비중은 %015.2f%%입니다.' % (a[-6:-1], b), end='\n\n')  # 우정렬 및 0채움

# python 3.0 이후 format 타입 : {지정변수:[빈공간채울문자][정렬방향<>][+양수부호][공간수][,][.소수자릿수]}
print('value is {{ {} }}'.format('%d' % b))
print('{}의 투자비중은 {}%입니다.'.format(a[7:12], '%15.2f' % b))
print('{name}의 투자비중은 {money}%입니다.'.format(name=a[7:12], money='%15.2f' % b))
print('{0}의 투자비중은 {1:<15.2f}%입니다.'.format('K씨', -1000.5))   # 좌정렬
print('{}의 투자비중은 {:^15.2f}%입니다.'.format('K씨', -1000.5))     # 가운데정렬
print('{}의 투자비중은 {:>15.2f}%입니다.'.format('K씨', -1000.5))     # 우정렬
print('{}의 투자비중은 {:0>15.2f}%입니다.'.format('K씨', -1000.5))    # 공백채우기
print('{}의 투자비중은 {:,.2f}%입니다.'.format('K씨', -1000.5))       # 천단위콤마표시
print('{0:+15,} 과 {1:+15,} 사이'.format(50000, -50000))             # 양수부호

# python 3.6 이후 f-string 타입
print(f'value is {{ {int(b)} }}')
print(f'{a[7:12]}의 투자비중은 {b:<15,.2f}%입니다.')
print(f'{a[7:12]}의 투자비중은 {b:^15,.2f}%입니다.')
print(f'{a[7:12]}의 투자비중은 {b:>15,.2f}%입니다.')


# -----------------------------
# List : list함수는 실행으로도 원본이 바뀜
# -----------------------------

# names = ['Lim', 'Kim', 'Park', 'Choi', 'Jung', 'Lee', 'Lim', 'Park']
movie = ['b', 'o', 'x']

# 개수--------------
len(movie)
movie.count('x')

# 찾기--------------
movie.index('x')  # 첫번째 인덱스 번호

# copy--------------
movie1 = movie           # 얕은 복사
movie2 = movie[:]        # 깊은 복사
movie3 = movie.copy()
import copy
movie4 = copy.deepcopy(movie)
print(id(movie), id(movie1), id(movie2), id(movie3), id(movie4), sep='\n')

# insert--------------
movie1.append(movie3)         # 마직막 위치에 원리스트(2단 list)로 삽입
movie2.extend(movie3)         # 마지막 위치에 원리스트를 풀어서 병합
movie3.insert(2, '위치:2')    # 원하는 인덱스 위치에 삽입
print(movie, movie1, movie2, movie3, sep='\n')    # movie1 과 같은 메모리의 movie 도 바뀜

# replace--------------
movie4 = movie3.copy()
movie3[0] = ['c', 'a', 'r']     # 원리스트(2단 list)로 삽입
movie4[0:1] = ['c', 'a', 'r']   # 원리스트를 풀어서 병합
print(movie3, movie4, sep='\n')

# delete--------------
movie = ['b', 'o', 'x']
del movie[0]             # 위치값으로 삭제
movie.remove('x')        # 원소값으로 삭제
movie[3:] = []           # 범위로 삭제
movie.pop()              # 마지막 위치값을 삭제하면서 삭제된 값을 반환
movie.pop(1)             # 해당 위치값을 삭제하면서 삭제된 값을 반환
movie.clear()            # 전부 비우기

# 리스트를 문자열로 변환--------------
lst = ['태연', 100, '정현', 200, '성진']
lst_string1 = ", ".join([str(a) for a in lst])
lst_string2 = ", ".join(map(str, lst))
print(lst_string1, lst_string2, sep="\n")

# 문자열을 리스트로 변환--------------
lst = '태연 | 진우 | 정현 | 하늘 | 성진'
lst_split = lst.split("|")
print(lst_split)

# 줄 단위로 구분된 문자열을 리스트로 변환--------------
lst = 'haha, \nhoho, \nhihi'
lst.splitlines()

# 반복문을 활용한 리스트 생성--------------
[i * 5 for i in range(1, 10) if i % 2 == 0]
[i * 5 if i % 2 == 0 else i for i in range(1, 10)]  # if-else문은 for문 앞에 위치**

# zip(): 리스트 또는 튜플을 묶어서 활용--------------
name = ['merona', 'gugucon']
price = [500, 1500]

to_list = list(zip(name, price))
to_dict = {a: b for a, b in zip(name, price)}

print(to_list, to_dict, sep="\n")
print(['%s의 가격은 %d입니다.' % (a, b) for a, b in zip(name, price)])
print(['%s의 가격은 %d입니다.' % (a, b) for a, b in to_list])
print(['%s의 가격은 %d입니다.' % (a, b) for a, b in to_dict.items()])

# enumerate(): 리스트를 딕셔너리로 변환--------------
names = ['태연', '진우', '정현', '하늘', '성진']
{'%d번' % (num): '%s' % (name) for num, name in enumerate(names, start=1)}
{'%d번' % (p[0]): '%s' % (p[1]) for p in enumerate(names, start=1)}


# -----------------------------
# Tuple 언팩킹
# -----------------------------

# 변수 앞에 *를 하나 붙여주면 지정되지 않은 모든 데이터를 담는다--------------
scores = (1, 2, 3, 4, 5, 6)
a, *b, c = scores  # 첫번째 데이터를 a, 마지막 데이터를 c로 바인딩하고 나머지를 *b에 바인딩
print(type(b), b, tuple(b), sep='\n')


def hap(num1, num2, num3, num4):
    return num1 + num2 + num3 + num4


scores = (1, 2, 3, 4)
print(hap(*scores))   # hap(scores[0], scores[1], scores[2], scores[3])


# 함수에서 여러 값을 리턴하면 튜플로 패킹되어 튜플 객체가 리턴된다--------------


def foo():
    return 1, 2, 3


print(type(foo()), foo(), sep='\n')


# -----------------------------
# Dictionary : 순서무시, index 무의미, []는 키값을 나타낼 뿐
# -----------------------------

aa = {1: 'one', 2: [1, 2, 3], '인사': '방가'}
bb = {'korea': 'seoul', 'japan': 'tokyo'}

# access--------------
list(aa.keys())
list(aa.values())

aa['인사']                # 키값이없으면 KeyError, 프로세스 중단
aa.get('인사')            # 키값이없으면 무반응, 다음 프로세스 진행
aa.get('gender', 'man')   # 키값이없을때 지정값 반환

# insert, update--------------
aa[1] = 'bye'
aa[6] = 'six'
aa.update(bb)       # 이어 붙이기
print(aa)

# delete--------------
del aa['인사']
aa.clear()

# 딕셔너리의 생성--------------
name = ['merona', 'gugucon', 'bibibig']
price = [500, 1000, 600]
{"메로나": 500, "구구콘": 1000}           # 일반 방법
dict(메로나=500, 구구콘=1000)             # 키가 문자열일 때 dict클래스에 담는 방법
dict(zip(name, price))                   # dict() + zip() 메서드
{k: v * 2 for k, v in zip(name, price)}    # 딕셔너리 컴프리헨션 방법: 연산 및 조건 식 가능
{k: v for k, v in zip(name, price) if v < 1000}


# -----------------------------
# 조건문, 반복문
# -----------------------------

# True  : 0이 아님, value 존재
# False : 0, [], None
# 논리연산자 : not, and ,or
# 비교연산자 : ==, !=, >=, <=

a = 257
b = 257

a == b   # == 값을 비교
a is b   # is 동일주소 여부 비교(파이썬은 메모리 절약을 위하여 정수256 초과부터 새메모리 할당)
id(a), id(b)  # 메모리 주소값

# if문: if:_[elif:]_[else:]--------------

# While문: while:_[break]_[continue]_[pass]_[else:]--------------

# for문 : for_in_:[else:]--------------


# =============================
# 파이썬 함수의 종류와 사용방법
# =============================

# 함수는 코드에 대한 이름표--------------
# ┌ 함수명으로 바로 사용: 빌트인 함수, 사용자 정의 함수
# └ OOO.함수명으로 사용: 다른 모듈(패키지) 함수, 클래스 함수

# 1. 빌트인 함수 (내장 함수)
#  print, open, int 등 파이썬 기본 함수

# 2. 사용자 정의 함수
#  def로 함수명, 입력인자, 반환값을 직접 정의해서 사용

# 3. 다른 모듈(패키지) 함수
#  다른 파이썬 파일에서 정의한 함수나 패키지의 함수를 사용
#  import [폴더명].[파일명].[함수명]으로 호출하여 사용

# 4. 클래스 함수
#  파이썬은 객체 지향 프로그래밍 언어로써, 모든 요소는 클래스로 되어 있고
#  클래스는 객체로 만들어져서 '객체명.함수명' 형태로 사용할 수 있다.

# 5. 패키지나 함수에 포함된 하위 옵션 확인
#  dir함수를 이용하면 '객체의 속성값과 함수'를 확인할 수 있다.
dir("a")       # 문자열객체가 사용 가능한 메서드
dir([10, 20])  # 리스트객체가 사용 가능한 메서드
dir(sys)       # sys모듈에서 사용 가능한 메서드


# -----------------------------
# return Method
# -----------------------------

# # 함수 속의 리턴은 결과값을 반환하며 함수를 중단한다
def f(treeHit):
    while treeHit < 10:
        treeHit = treeHit +1
        print("나무를 %d번 찍었습니다." % treeHit)
        if treeHit == 5:
            return "나무 넘어갑니다."
    return -1
f(9)

# 가변 매개변수--------------

# 인자값의 개수를 가변정의하기 위해서는 함수 파라미터 앞에 *를 붙여주어 튜플 객체로 저장
# 가변 매개변수는 하나만 사용할 수 있으며, 가변 매개변수 뒤에는 일반 매개변수가 올 수 없다.

def print_n_times(n, *values):
    for i in range(n):
        for i in values:
            print(i)
        print()


print_n_times(3, "안녕하세요.", "즐거운", "파이썬 코딩이예요.")

# 리스트/튜플를 가변 매개변수로 전달하기


def foo(a, b, c):
    print(a, b, c)


data = [1, 2, 3]
foo(*data)


# 기본 매개변수--------------

# 매개변수=값의 형태로 입력, 매개변수를 입력하지 않으면 기본값이 설정,
# 기본 매개변수 뒤에는 일반 매개변수가 올 수 없다.
def print_n_times(value, n=2):
    for i in range(n):
        print(value)


print_n_times("안녕하세요.")  # n이 입력되지 않았으므로 n=2

# 기본 매개변수가 가변 매개변수보다 앞에 올 때는 기본 매개변수의 의미가 없어짐.


def print_n_times(n=2, *values):
    for i in range(n):
        print(values)
        print()


print_n_times("안녕하세요.", "즐거운", "파이썬")      # n="안녕하세요"

# 가변 매개변수가 기본 매개변수보다 앞에 올 때는 기본 매개변수 생존.
# 함수호출할때 기본변수를 지정하면 가변/기본매개변수를 함께 사용 가능


def print_n_times(*values, n=2):
    for i in range(n):
        print(values)
        print()


print_n_times("안녕하세요.", "즐거운", "파이썬", 3)   # n=2 (3이 아님)
print_n_times("안녕하세요.", "즐거운", "파이썬", n=3)  # n=3


# 키워드 매개변수: 매개변수 이름을 지정해서 호출--------------

# 함수 파라미터 앞에 **를 붙여주며, 키워드와 값이 딕셔너리객체로 저장됨
def foo(**kwargs):
    print(kwargs)


foo(a=1, b=2, c=3)


# 가변매개변수와 키워드매개변수를 함께 사용하면 결과는 튜플과 딕셔너리로 분리됨--------------

def foo(*args, **kwargs):
    print(args)       # 튜플객체 반환
    print(kwargs)     # 딕셔너리객체 반환


foo(1, 2, 3, a=1, b=1, c=2)


# lambda함수--------------
# 일회용함수, lambda 다음에 적어주는 x가 입력을 의미하고 콜론(:) 이후의 표현식이 함수의 기능이며 이 값을 리턴

a = lambda x: 5 * x   # a가 람다함수 바인딩
a(5)                  # a()를 통해 함수 호출


# -----------------------------
# yield Method (제너레이터 함수)
# -----------------------------

# 제너레이터 함수: return 키워드 대신 yield라는 키워드를 사용하는 함수
# return 이 만든 빵을 전달한 후 다시 빵을 만들 때 다시 옷을 갈아입고 손을 씻는다면,
# yield 는 만든 빵을 전달한 후(코드 중지 상태를 저장) 즉시 빵을 만드는 것

# yield가 사용된 함수는 ( )를 붙여도 코드가 바로 실행되지 않는다.
# 제너레이터 객체를 생성되고 코드가 실행될 준비만 합니다.
# next( ) 함수 호출을 통해 제너레이터 객체 내의 코드를 실행하여 값을 호출부로 리턴하고
# 현재 상태를 그대로 유지합니다.
# 다시 호출부에서 next() 함수를 호출하면 앞서 중지된 상태부터 코드가 실행됩니다.

def num_gen():
    for i in range(3):
        yield i


g = num_gen()       # 제너레이터 객체 생성
num1 = next(g)
num2 = next(g)
num3 = next(g)
print(num1, num2, num3)

# next(g)를 사용하지 않고 제너레이터로부터 값을 가져오는 더 쉬운 방법은
# for 문을 사용하는 것
# (내부적으로 next(g)를 호출하여 StopIteration이 발생할 때 까지 진행)


def num_gen():
    for i in range(3):
        yield i


g = num_gen()
for i in g:
    print(i)


def 빵만들기(n):
    # 빵쟁반 = []
    for i in range(n):
        빵 = "빵" + str(i)        # 빵0, 빵1, ..., 빵99
        # 빵쟁반.append(빵)
    # return 빵쟁반
        yield 빵  # 주의!! 들여쓰기


def 빵포장(빵):
    print("{} 포장완료".format(빵))


for i in 빵만들기(100):
    빵포장(i)

# 리스트에 들어있는 요소를 한 개씩 바깥으로 전달


def number_generator():
    x = [1, 2, 3]
    yield from x


k = number_generator()
num1 = next(k)
num2 = next(k)
num3 = next(k)
print(num1, num2, num3)


# -----------------------------
# 데코레이터와 클로저
# -----------------------------

# 데코레이터: 다른 함수를 장식, 함수를 수정하지 않은 상태에서 추가 기능을 구현할 때 유용
# 클로저: 외부 함수 안에 내부 함수를 정의하고 정의된 내부 함수를 리턴하는 구조

def trace(func):                            # 호출할 함수를 매개변수로 받음
    def wrapper():                          # 호출할 함수를 감싸는 함수
        print(func.__name__, '함수 시작')    # __name__으로 함수 이름 출력
        func()                              # 매개변수로 받은 함수를 호출
        print(func.__name__, '함수 끝')
    return wrapper                          # wrapper 함수 반환


def hello():
    print('hello')


trace_hello = trace(hello)    # 데코레이터에 호출할 함수를 넣음
trace_hello()                 # 반환된 함수를 호출


# @로 데코레이터 사용하기--------------

def trace(func):                            # 호출할 함수를 매개변수로 받음
    def wrapper():
        print(func.__name__, '함수 시작')    # __name__으로 함수 이름 출력
        func()                              # 매개변수로 받은 함수를 호출
        print(func.__name__, '함수 끝')
    return wrapper                          # wrapper 함수 반환


@trace    # @데코레이터
def hello():
    print('hello')


hello()    # 함수를 그대로 호출


# -----------------------------
# Class
# -----------------------------

# 클래스는 데이터와 이를 처리하는 메서드(함수)로 구성--------------

class Person:
    pass


p = Person()  # 해당 클래스의 객체를 생성, 클래스와 객체는 각자의 메모리 공간을 가짐


# 객체 이름 공간에 변수 생성--------------

class Person:
    pass


p = Person()
p.data = 3
# 점(.)의 의미는 p 공간안을 의미하며 p가 가리키는 공간에
# {"data":3} 딕셔너리를 객체 공간에 저장


# 클래스 공간에 데이터 저장하기--------------

class Person:
    data = 4   # 클래스공간에 {"data":4} 를 저장


class MyClass:
    count = 0

    def __init__(self):    # 생성자: 객체생성될때 자동실행(객체 초기화 활용)
        MyClass.count += 1

    def get_count(self):
        return MyClass.count


a = MyClass()
b = MyClass()
c = MyClass()

print(a.get_count())
print(MyClass.count)


# -*- coding: utf-8 -*-

class Man:

    # 클래스 전역변수
    cnt = 0                       # 전역변수: "클래스명.변수명"으로 사용
    class_var = 1000
    class_list_var = [2000]

    def __init__(self, name):
        Man.cnt += 1
        self.name = name           # 지역변수: "self.변수명"으로 사용
        self.init_var = 3000
        self.init_list_var = [4000]
        print('({}이 생성되었습니다.)'.format(self.name))

    def die(self):
        Man.cnt -= 1
        print('{}는 죽었습니다.!!!'.format(self.name))
        if Man.cnt == 0:
            print('{}는 최후의 생존자였습니다'.format(self.name))
        else:
            print('아직 {:d}명의 생존자가 남았습니다'.format(Man.cnt))

    @classmethod
    # class의 메서드(함수)는 인스턴스만 호출 할 수 있으나,
    # @classmethod를 사용하면 자기 매서드를 호출 할 수 있다.
    def how_many(how):
        print('현재 {}명이 남았습니다'.format(Man.cnt))

# -----------------------------------------------------


# 인스턴스 생성
Actor1 = Man('맨1')
Actor2 = Man('맨2')

# 인스턴스(실체)는 메소드를 호출한다
Actor1.die()
Actor2.die()
Actor3 = Man('맨3')

# 클래스(가상)는 메서드를 호출 할 수 없다(TypeError)
Man.die()

# 클래스는 @classmethod로 정의된 메서드는 호출 할 수 있다
Man.how_many()

# 인스턴스는 변수를 호출한다
print(f'유닛이름 : {Actor1.name}, 값 : {Actor1.class_var}, {Actor1.init_var}')

# 인스턴스는 각자의 메모리 영역에 클래스변수를 저장
print(Actor1.class_var, Actor2.class_var)
Actor1.class_var = 5000
print(Actor1.class_var, Actor2.class_var)

# [예외] "리스트" 클래스변수는 같은 메모리 영역을 공유
print(Actor1.class_list_var[0], Actor2.class_list_var[0])
Actor1.class_list_var[0] = 6000
print(Actor1.class_list_var[0], Actor2.class_list_var[0])

# -----------------------------------------------------
# @property를 붙여주면 속성처럼 사용할 수 있음(함수호출을 위해 ()를 붙일 필요 없음)


class Car:

    def __init__(self, model):
        self.model = model

    @property
    def get_model(self):
        return self.model


c = Car("GV80")
print(c.get_model())        # c.get_model()이 아님

c.model

# -----------------------------------------------------
# 삼각김밥 클래스


class Samgak:
    def __init__(self):
        self.source = '기본소스'
        self.kim = '광천김'
        self.bab = '쌀밥'
        self.food = '야채'

    def set_source(self, source_name):
        self.source = source_name

    def change_kim(self, kim_name):
        self.kim = kim_name

    def change_bab(self, bab_name):
        self.bab = bab_name

    def set_food(self, food_name):
        self.food = food_name

    def print(self):
        s1 = 'BlockDMask 가 맛있는 삼각김밥을 만들었습니다.\n'
        s1 += f'김은 {self.kim} 입니다.\n'
        s1 += f'밥은 {self.bab} 사용 하였고\n'
        s1 += f'소스는 {self.source} 촵촵 뿌리고\n'
        s1 += f'메인은 {self.food}을 넣었습니다.\n'
        print(s1)


# 참치 삼각김밥 인스턴스
chamchi = Samgak()
chamchi.print()
chamchi.set_food('동원참치')
chamchi.set_source('마요네즈')
chamchi.print()

chamchi.source
del chamchi.source
del chamchi

# -----------------------------------------------------
# inheritance(상속)
# 상속은 기존 클래스를 변경하지 않고 기능을 추가하거나 기존 기능을 변경 할 때 사용.
# 기존 클래스를 수정/추가 하면 되지만, 기존 클래스가 라이브러리 형태로 제공되거나
# 수정이 허용되지 않는 상황이라면 상속을 사용해야 함


class Animal():         # 부모 클래스
    def walk(self):
        print('걷는다')

    def eat(self):
        print('먹는다')

    def greet(self):
        print('인사한다')


class Human(Animal):    # 자식 클래스
    def wave(self):
        print('손을 흔든다')

    def greet(self):    # 자식 클래스의 greet()가 우선
        self.wave()
        super().greet()  # super(): 부모 클래스의 greet()를 호출


person = Human()
person.walk()
person.eat()
person.wave()
person.greet()


# import 동작할 때는 미사용
if __name__ == "__main__":
    print(PI)
    a = Math()
    print(a.solv(2))
    print(sum(PI, 4.4))
    r1, r2 = root(100, 5, 20)
    print(r1, r2)


# -----------------------------
# 유닛 테스트
# -----------------------------

# mymath.py 파일 저장 ----
def average(a):
    return sum(a) / len(a)


# mymath.py 파일 저장 ----
import pytest
from mymath import *


def test_average():
    assert average([1, 2, 3]) == 1


"""
# 터미널에서 테스트 실행
1. 파일명에 test_(접두사)/_test(접미사)의 파일명으로 저장
2. 터미널에서 pytest라고 입력
3. 자동으로 접두/접미사가 붙은 파일을 검사하여 결과를 화면에 출력
   - 파일명 뒤에 .이 있는 것은 테스트 통과를 의미
   - 파일명 뒤에 F가 있는 것은 테스트 실패를 의미
"""


# -----------------------------
# 로깅 및 에러
# -----------------------------

# test.py 파일 저장 ----
def hap(a, b):
    ret = a + b
    print(a, b, ret)     # 함수의 입/출력 확인을 위한 print 구문
    return ret


result = hap(3, 4)


# 함수 결과값 확인 방법(1) print()--------------
print(a, b, ret)


# 함수 결과값 확인 방법(2) 터미널에서 결과값 파일 저장--------------
# python test.py > log.txt


# 함수 결과값 확인 방법(3) logging 모듈 + 터미널 실행으로 결과값 출력--------------
# print 대신에 logging.info() 함수를 사용하는 겁니다. 이때 info 함수의 인자는 문자열 타입이어야 합니다.

# run.py 파일 저장 ----
import logging

# logging.basicConfig(level=logging.DEBUG)  # logging 레벨 설정
logging.basicConfig(filename="mylog.txt", level=logging.INFO)  # 파일로 결과 저장


def hap(a, b):
    ret = a + b
    logging.info(f"input: {a} {b}, output={ret}")
    return ret


result = hap(3, 4)
# -----------------------
# logging 레벨
# DEBUG		상세한 정보를 출력
# INFO		예상대로 작동하는지를 확인
# WARNING	소프트웨어는 정상 동작하는데 예상치 못한 일이 발생한 것에 대해 표시
# ERROR		소프트웨어의 일부가 정상적으로 동작하지 않는 경우에 대해 표시
# CRITICAL	심각한 에러 상황에 대해 표시


# -----------------------------
# 예외 처리
# -----------------------------

# 기본--------------

'''
try:
except:
else:
finally:
'''

classrooms = {'1반': [172, 188, 197, 189], '2반': [180, 191]}

try:
    for class_id, heights in classrooms.items():
        for height in heights:
            if height > 190:
                print(class_id, '에 190이 넘는 학생이 있습니다')
                raise StopIteration           # 강제로 에러발생시켜서 종료처리
except StopIteration:                         # 발생에러를 정의
    pass                                      # 에러처리 않고 통과
else:                                         # 정의되지않은 모든에러처리
    print('알 수 없는 오류가 발생 되었습니다')
finally:                                      # 어떤 상황에서도 항상 수행(file.close()등)
    print('예외 발생해도 무조건 수행 됩니다')

# exception: 모든 종류의 '에러메세지'를 반환--------------
try:
    list = []
    print(list[0])
except Exception as err:
    print('오류: ', err)

# traceback 내장모듈: 모든 종류의 '에러메세지 + 에러발생위치'를 반환--------------
try:
    list = []
    print(list[0])
except Exception:
    import traceback
    traceback.print_exc()

# 사용자 정의 예외--------------
# 사용자 정의 예외는 기본적으로 Exception 클래스를 상속 받아서 사용합니다.


class MyError(Exception):
    """Base 에러 클래스"""
    pass


class CreateAskError(MyError):
    def __str__(self):
        return "정보가 올바르지 않습니다."


class InsufficientFundsAsk(MyError):
    def __str__(self):
        return "잔고가 부족합니다."


# -----------------------------
# 동시성 프로그래밍
# -----------------------------

# 동시성과 병렬성--------------
# 동시성(멀티태스크): 동시에 하는 것처럼 보이도록 순간순차적으로 작업 처리
# 병렬성(멀티코어): 멀티코어를 활용하여 실제로 병렬 작업 처리


# 파이썬 스레드--------------

import threading
import time


class Worker(threading.Thread):
    def __init__(self, name):
        super().__init__()
        self.name = name            # thread 이름 지정

    def run(self):
        print("sub thread start ", threading.currentThread().getName())
        time.sleep(3)
        print("sub thread end ", threading.currentThread().getName())


print("main thread start")

for i in range(5):
    name = "thread {}".format(i)
    t = Worker(name)                # sub thread 생성
    # t.daemon = True                 # 서브 스레드들을 데몬스레드(메인스래드가 종료하면 작업완료여부와 관계없이 종료)로 만듦
    t.start()                       # sub thread의 run 메서드를 호출

print("main thread end")


# Fork와 Join: 모든 스레드 작업 완료 확인 후 후반 작업이 필요한 경우에 적용--------------

# Fork : 메인스레드에서 서브스레드가 분리되는 기점
# Join : 분할된 스레드가 모두 완료되는 기점

import threading
import time


class Worker(threading.Thread):
    def __init__(self, name):
        super().__init__()
        self.name = name            # thread 이름 지정

    def run(self):
        print("sub thread start ", threading.currentThread().getName())
        time.sleep(5)
        print("sub thread end ", threading.currentThread().getName())


print("main thread start")

t1 = Worker("1")        # sub thread 생성
t1.start()              # sub thread의 run 메서드를 호출

t2 = Worker("2")        # sub thread 생성
t2.start()              # sub thread의 run 메서드를 호출

t1.join()
t2.join()

print("main thread end")


# -----------------------------
# 다양한 모듈
# -----------------------------

# [검색 방법]
# pypi(python package index)로 패키지모듈 검색
# list of python builtins로 내장함수 검색(Docs>Library Reference)
# list of python modules로 외장함수 검색(Docs>Module Index)


# ################
# re모듈
# ################

# # 1. re 모듈의 메소드 ----------

# 종류       기능                    찾는경우                없는경우
# match      문자열 첫글자부터 검색   match object 1개       None
# search     문자열 전체를 검색       match object 1개       None
# findall    문자열 전체를 검색       문자열 리스트          빈 리스트
# finditer   문자열 전체를 검색       match object iterator  None
# fullmatch  패턴과 문자열이 완벽일치 match object 1개       None
# sub        일종의 replace
# compile    정규식을 반복 사용


# # 2. match object의 메소드 ----------

# 종류   기능                           예시
# group  매칭된 문자열을 반환            people
# ※ groups(), group(int)  소괄호()가 존재하는 패턴에서만 활용
# start  매칭된 문자열의 시작 위치       5
# end    매칭된 문자열의 끝 위치         11
# span   매칭된 문자열의(시작, 끝)튜플   (5, 11)


# # 3. 정규표현식 ----------

# # Character classes
# .	        any character except newline
# \w\d\s	any word(알파벳+숫자+언더스코어), digit, whitespace([ \t\n\r\f\v]와 동일, 빈칸은 공백문자)
# \W\D\S	not any word, digit, whitespace
# abc       문자열('abc')를 찾음
# [abc]	    any of a, b, or c
# [^abc]	not any of a, b, or c
# [0-9]                character between 0 & 9
# [a-zA-Z]             알파벳
# [ㄱ-ㅎ|ㅏ-ㅣ|가-힣]   한글

# # Anchors
# ^abc$	    start / end of the string
# \b\B	    word, not-word boundary
# \b(?!\bto\b)\w+\b  ' to '단어 제외('too'는 포함)

# # Escaped characters
# \.\*\\	escaped special characters
# \t\v\n\r	tab, vertical tab, linefeed, carriage return

# # Groups & Lookaround
# (abc)	    capture group
# \1	    backreference to group #1(그룹1 표현식에서 캡쳐한 결과물을 복사)
# (?:abc)	non-capturing group(캡쳐그룹 해제)
# (?=abc)	positive lookahead, 조건 앞의 문자가 조건을 만족하는 경우 그것만 선택
# (?!abc)	negative lookahead, 조건 앞의 문자가 조건을 만족하는 경우 그것만 제외하고 선택
# (?<=abc)	positive lookbehind, 조건 뒤의 문자가 조건을 만족하는 경우 그것만 선택
# (?<!abc)	negative lookbehind, 조건 뒤의 문자가 조건을 만족하는 경우 그것만 제외하고 선택

# # Quantifiers & Alternation
# a*a+a?	  0 or more, 1 or more, 0 or 1
# a{5}a{2,}	  exactly five, two or more
# a{1,3}	  between one & three
# a+?a{2,}?	  match as few as possible(1 if 1 or more, two if two or more)
# ab|cd	      match ab or cd


# # 4. 예제 ----------

# 4-1. re 모듈의 메소드 예제

import re
str = 'love people around you, love your work, love yourself'

# match: 문자열의 첫글자부터 검색(결과: 처음 1개의 match 객체만 반환)
result = re.match('ove', str)
print(result)

# search: 문자열의 전체를 검색(결과: 처음 1개의 match 객체만 반환)
result = re.search('ove', str)
print(result)

# findall: 문자열의 전체를 검색(결과: 전부 찾아서 리스트로 반환)
result = re.findall('love', str)
print(result)

# finditer: 문자열의 전체를 검색(결과: 전부 찾아서 iterator(순회객체)로 반환)
result = re.finditer('love', str)
for iter in result:
    print(iter)

# fullmatch: 패턴과 문자열이 완벽하게 일치할 때 반환
result = re.fullmatch('love people around you, love your work, love yourself', str)
print(result)

result = re.fullmatch('.*', str)
print(result)

# sub(패턴, 교체할 문자열, 적용 문자열, 교체횟수)
str2 = '010-2343-9537'
result = re.sub(r'(?<=\d{3}-\d{4}-)\d{4}', '****', str2)
print(result)

re.sub("2343", "1234", str2)
re.sub(r"[^\d]", "", str2, 1)

# compile(패턴, 플래그): 정규식을 반복 사용할 때 지정 필요
c = re.compile('c')
print(c.sub('zzz', 'abcdefg'))
print(c.search('vcxdfsa'))

# 4-2. match 객체의 메소드 예제

import re
str = 'love people around you, love your work, love yourself'
result = re.search('people', str)

# group(): 매칭된 문자열 반환
print(result.group())

# start(): 매칭된 문자열의 시작 위치 반환
print(result.start())

# end(): 매칭된 문자열의 끝 위치 반환
print(result.end())

# span(): 매칭된 문자열의 (시작, 끝) 위치 튜플 반환
print(result.span())

# groups(), group(int)
result = re.match(r'(\d{2})-(\d{3,4})-(\d{4})', '02-123-1234')
print(result.group())
print(result.groups())
print(result.group(0))
print(result.group(1))
print(result.group(2))


# ################
# 유용한 모듈
# ################

# eval()함수:  수식 문자열을 실행하는 함수--------------
ex = '5+4'
eval(ex)

# 동적 변수--------------
globals()['test'] = 100	 # {'test': 100} 글로벌 변수에 추가
print(test)

for i in range(1, 4):
    locals()['test' + str(i)] = i * 10   # {생략, 'test1': 10, 'test2': 20, 'test3': 30}
    # locals()['test{}'.format(i)] = i * 10
print(test1, test2, test3)

print(globals())
print(locals())

# 무작위수--------------
import random
import numpy as np
np.random.choice(['가위', '바위', '보'], 10)  # 랜덤 쵸이스
random.choice(['가위', '바위', '보'])         # 랜덤 쵸이스
random.random()                              # 0.0~1.0사이의 실수 값을 난수로 발생
random.randint(10, 50)                       # 10과 50사이의 정수 값을 난수로 발생

# 운영체제 명령어--------------
import os
ds_path = os.getcwd()           # 현재 경로
ds_file_path = __file__         # 현재 파일의 경로(파일명 포함)
print(ds_path, ds_file_path, sep='\n')

os.path.abspath(ds_file_path)   # 절대경로로 변환
os.path.dirname(ds_file_path)   # 마지막 경로,파일명을 제외하고 반환
os.path.basename(ds_file_path)  # 마지막 경로,파일명을 반환
os.path.split(ds_file_path)     # 마지막 경로,파일명을 분리 반환
os.path.exists(ds_file_path)    # 파일 또는 경로가 존재 여부 반환

mk_dir_name = os.path.join(ds_path + os.sep + 'Test')   # 경로 병합
os.mkdir(mk_dir_name)           # 새로운 폴더를 만들기
os.chdir(mk_dir_name)           # 작업 경로 변경
print(os.getcwd())

for i in range(6):
    file_name = 'test' + str(i) + '.txt'
    f = open(file_name, 'w', encoding='utf8')
    f.close()

os.listdir(mk_dir_name)                 # 경로 안의 하위경로명 및 파일명 리스트
os.rename('test0.txt', 'test6.txt')     # 파일명 또는 경로명 바꾸기
os.remove('test1.txt')                  # 파일삭제
os.system('copy test3.txt test7.txt')   # 시스템의 유틸리티나 dos명령어 사용
os.system('del *.txt')
os.chdir(os.path.dirname(mk_dir_name))
print(os.getcwd())
os.rmdir(mk_dir_name)                   # 폴더 삭제

# datetime--------------
import datetime
current = datetime.datetime(2020, 5, 16, 23, 10, 1, 100)
current_time = current.replace(microsecond=0)
currentdate = datetime.datetime.now().date()
currentdate2 = datetime.date.today()
currenttime = datetime.datetime.now().time()
combine_datetime = datetime.datetime.combine(currentdate, currenttime)
timestamp = datetime.datetime.now().timestamp()
fts = datetime.datetime.fromtimestamp(timestamp)
chg = datetime.datetime.strftime(current, '%b %d %y %H %M %S')
chg2 = current.strftime('%b %d %y %H %M %S')
todatetime = datetime.datetime.strptime(chg2, '%b %d %y %H %M %S')
days_cap = (combine_datetime - current).days
hours_cap = (combine_datetime - current).seconds / 3600
mins_cap = (combine_datetime - current).seconds / 60
scds_cap = (combine_datetime - current).seconds
base_day = datetime.timedelta(days=3)
previous_day = current - base_day
next_day = current + base_day

print(datetime.datetime.__doc__)
print('-' * 40)
print("날짜와 시간 출력: ", current)
print("마이크로세컨드 제외: ", current_time)
print("날짜 출력: ", currentdate)
print("시간 출력: ", currenttime)
print("날짜와 시간 합치기: ", combine_datetime)
print('-' * 40)
print("timestamp는 float형: ", timestamp)
print("timestamp를 datetime형으로 변환: ", fts)
print('-' * 40)
print("datetime을 원하는 형식의 str로 변환2: ", chg)
print("datetime을 원하는 형식의 str로 변환2: ", chg2)
print("str을 datetime으로 변환: ", todatetime)
print('-' * 40)
print("날짜 차이: ", days_cap)
print("시간 차이: ", hours_cap)
print("분 차이: ", mins_cap)
print("초 차이: ", scds_cap)
print('-' * 40)
print("날짜계산과 시간계산을 위해서는 timedelta클래스를 사용: ", base_day)
print("이전날짜 -> ", previous_day)
print("다음날짜 -> ", next_day)


# -----------------------------
# 데이터 입출력
# -----------------------------

# r  읽기
# w  쓰기: 파일 만들고, 기존 파일 있으면 덮어쓰기
# x  쓰기: 파일 만들고, 기존 파일 있으면 오류 발생
# a  붙여쓰기: 기존 파일 내용 뒤에 추가하고, 없으면 새로 만듦
# r+ 붙여쓰기: 기존 파일 내용 앞에 추가
# b  바이너리모드로 열기


# 텍스트 입출력--------------

# 읽기
f = open('test.txt', 'r', encoding='utf8')
print(f.read())            # 결과값 str
f.close()
# print(f.readlines())     # 한 번에 읽기(결과값 list)
# print(f.readline())      # 한 줄씩 읽기(while 문과 함께 사용)
# with open('matrix_quotes.txt', 'r') as text_file:
#     matrix_quotes_list = []
#     line = text_file.readline()
#         while line != '':
#             matrix_quotes_list.append(line)
#             line = text_file.readline()

# 텍스트 모드에서 인코딩, 디코딩 에러 발생 시 처리 : errors
# open('file.txt', 'rw', errors = 'ignore')
# 'ignore' 에러 무시
# 'backslashreplace' 지원되지 않는 문자를 \로 시작되는 escape sequence로 바꿔서 기록

import csv

with open('listfile.csv', 'r', encoding='utf-8') as f:
    rdr = csv.reader(f)
    for line in rdr:
        print(line)

with open('listfile.csv', 'r', encoding='utf-8') as f:
    rdr = csv.reader(f)
    for i, line in enumerate(rdr):
        if i == 0:
            date = line
        elif i == 1:
            val1 = line
        elif i == 2:
            val2 = line

# 쓰기
f = open('test.txt', 'w', encoding='utf8')   # 'cp949' for Windows
data = '간단한 파일쓰기'
f.write(data)
f.close()
# writelines() : 한 번에 쓰기(list유형)
# write()      : 한 줄씩 쓰기(for loop 문과 함께 사용)
# with open('matrix_quotes.txt', 'w') as f:
#     for line in matrix_quotes:
#         f.write(line)

# 자동으로 파일을 닫는 with문
with open('test2.txt', 'w', encoding='utf8') as f:
    f.write('간단한 with문 쓰기')

# single list 저장
import csv
date = ['200801', '200802', '200803', '200804', '200805']
val1 = [50, 100, 300, 150, 200]
val2 = [2000.0, 2200.5, 2320.3, 2250.0, 2200.0]

with open('listfile.csv', 'w', newline='') as f:  # newline 설정을 안하면 한줄마다 공백있는 줄이 생긴다.
    writer = csv.writer(f)
    writer.writerow(date)
    writer.writerow(val1)
    writer.writerow(val2)

# multi list 저장
import csv
listitem = [["Korea", "USA", "Japan"], ["Korea1", "USA1", "Japan1"], ["Korea2", "USA2", "Japan2"]]

with open('List_csv1.csv', 'w', encoding='utf-8', newline='') as file:
    write = csv.writer(file)
    write.writerows(listitem)


# 화면 출력을 파일로 저장하기--------------

# 방법1. 특정 출력 결과를 파일 저장: print 옵션
import sys
f = open('output.txt', 'w')
print('파일로 저장됩니다', file=f)
f.close()

# 방법2. 모든 출력 결과를 파일 저장: sys.stdout
import sys
temp = sys.stdout
sys.stdout = open('output.txt', 'w')  # 이어쓰기 옵션 'a'
for i in range(10):
    print(i)
sys.stdout.close()
sys.stdout = temp

# 방법2-1. with문
import sys
temp = sys.stdout
with open("output.txt", "w") as sys.stdout:
    for i in range(10):
        print(i)
    sys.stdout = temp

# 방법3. 명령 프롬프트에서 결과를 파일 저장
# python test.py > output.txt    # 이어쓰기 옵션 >>


# 이미지로 저장--------------
import dataframe_image as dfi
dfi.export(df, '파일명.jpg', max_rows=-1)    # -1은 모든 행을 의미
dfi.export(df, '파일명15.jpg', max_rows=15)  # 행 개수를 15개로 제한(줄임 표시)


# 바이너리 입출력--------------

import pickle
fp = open('pickle_dict.pickle', 'wb')  # 바이너리임을 알수있도록 w뒤에 b를 붙인다
dic = {'name': 'Park', 'age': 50, 'hobby': ['soccer', 'baseball', 'basket']}
pickle.dump(dic, fp)                  # 객체를 파일에 저장
fp.close()

fr = open('pickle_dict.pickle', 'rb')  # 바이너리임을 알수있도록 r뒤에 b를 붙인다
load = pickle.load(fr)                # 파일에서 객체를 불러오기
print(load)
fr.close()

with open('pickle_dict.pickle', 'rb') as fp:
    print(pickle.load(fp))


# 엑셀 입출력--------------

import openpyxl

# 파일 읽기
wb = openpyxl.load_workbook(r'D:\DeskTop\ctiger\Dropbox\Goodjob\Pi\trainingdata\sample_dir\stats.xlsx')
ws = wb.get_sheet_by_name('people')
# ws = wb.worksheets[0]

# ## open multi Worksheets
# for i, sheetname in enumerate(wb.sheetnames):
#     locals()["ws{}".format(i)] = wb[sheetname]

# 내용 접근
ws.cell(row=8, column=9).value  # 1부터 시작
ws['A3:C3']
ws['A3':'C3']
ws.max_column
ws.max_row
ws.dimensions[1]
ws.rows
ws.columns
i = 1
print(int(ws[str(chr(i + 65)) + "4"].value))  # B4
# 대문자 A ~ Z 는 chr(65) ~ chr(90), 소문자 a ~ z 는 chr(97) ~ chr(122)

# 내용 쓰기
ws['A1'] = "바꿔바꿔 세상을 다바꿔"
ws.cell(row=4, column=2, value='대한민국')

# 내용 삭제
ws.delete_rows(29, 73)
ws.delete_cols(8, 10)

# 행,열 삽입,삭제하기
ws.insert_rows(2)   # 두번 째 행에 삽입
ws.insert_cols(2)   # 두번 째 열에 삽입
ws.delete_rows(2)
ws.delete_cols(2)

# 파일 저장
wb.save(r'C:\Drive\DeskTop\ctiger\Dropbox\Goodjob\Pi\_common\sample_dir\stats_adj.xlsx')

# Tip. Cell Style Collection
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import Color, PatternFill

# 글꼴
font_15 = Font(name='맑은 고딕', size=15, bold=True)

# 정렬
align_center = Alignment(horizontal='center', vertical='center')
align_vcenter = Alignment(vertical='center')

# 테두리
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# 셀 색상
fill_orange = PatternFill(patternType='solid', fgColor=Color('FFC000'))

# 스타일 설정
for row in ws['B2:J2']:
    for cell in row:
        cell.font = font_15
        cell.alignment = align_center
        cell.fill = fill_orange
        cell.number_format = '#,##0.00%'


#######################################
# 필터링
#######################################

# 불린 인덱싱: &,|,~(데이터에서의 not),^(xor)
df[~df.건물.str.contains('(KR|수퍼)', na=False)]  # 포함하지 않는 것1, (결측치가 있는 경우 'na=False'
df[df.건물.str.contains('(KR|수퍼)') is False]    # 포함하지 않는 것2
mask1 = (df2.나이 < 16) & (df2.성별 == '여')
mask2 = (df2.나이 < 16) | (df2.성별 == '남')
df2.loc[mask1, :]
df2.loc[mask2, ['나이', '성별', '학교명']]
df_3.loc[df_3['인구'] > 50000, ['지역', '인구']]
df_3[df_3.인구 > 50000]
df_3[df_3['지역'] != '서울']
df_3[pd.isnull(df_3['기온'])]
df_3[pd.notnull(df_3['기온'])]
#  NaN이 아닌 값을 구하려면 notnull 메서드를 사용
ds[ds.notnull()]

# isin() 메소드 활용
isin_filter = df2['년'].isin(['95', '96', '97'])  # isin(): 해당 값 유무에 따른 불린값 반환
df2[isin_filter]

# 기타
df_3.sample(frac=0.5)  # ramdom(%)
df_3.sample(n=3)       # ramdom(n)
df_3.nlargest(2, '기온')
df_3.nsmallest(2, '기온')


#######################################
# 전처리
#######################################

import pandas as pd
import numpy as np

tdf = pd.DataFrame({'id': ['A_부산', 'A_세종', 'C_광주', 'B_제주', 'C_서울', 'A_강릉'],
                    'pop': ['2,600', '1,500', '4,000', '900', '2,600', '3,300'],
                    'name': ['Alice', 'Bob', 'Charlie', 'Dave', 'Ellen', 'Frank'],
                    'age': [24, 42, 18, 68, 24, 30],
                    'state': ['NY', 'CA', 'CA', 'TX', 'CA', 'NY'],
                    'point': [64, 24, 70, 70, 88, 57]})

print(tdf, tdf.dtypes, sep='\n')

# .str.count: 원소에서 특정 문자 갯수 구하기
tdf[tdf.id.str.count('울|종') != 0]

# str.contains: 조건에 만족하는 문자의 포함여부 반환
tdf[tdf['id'].str.contains('울|종', na=False)]  # 결측치가 있는 경우 'na=False'
tdf.state.str.contains('CA', na=False)
len(tdf[tdf.state.str.contains('CA', regex=True)])

# 멤버쉽연산자: 검색한 문자열 포함 여부 반환
# 패턴 in 텍스트, 패턴 not in 텍스트

# find, index 계열 함수: 검색한 문자열 위치 반환
# str.find(sub, start, end)    # 처음찾은인덱스 또는 -1
# str.rfind(sub, start, end)   # 마지막찾은인덱스 또는 -1
# str.index(sub, start, end)   # 처음찾은인덱스 또는 ValueError
# str.rindex(sub, start, end)  # 마지막찾은인덱스 또는 ValueError

# with 계열 함수: 패턴이 문자열의 시작 또는 끝에 포함되어 있는지 여부 반환
# str.startswith(prefix, start, end)
# str.endswith(suffix, start, end)

# str.findall: 원소에서 정규표현식을 만족하는 모든 문자 추출
tdf.name.str.findall(r'[aeiouAEIOU]+')

# str.extract: 중괄호()를 사용하여 추출한 문자로 신규 열을 생성
tdf.name.str.extract(r'([aeiouAEIOU]+)')      # 정규표현식을 만족하는 첫번째 원소 추출
tdf.name.str.extract(r'(.+)(li)(\w)*')        # 3개그룹으로 나누면 3개열을 생성
tdf.age.astype(str).str.extract('(\d(?=8))')  # str형 변환

# .str.split(slice): 특정문자 또는 인덱스번호로 분리
tdf['class'] = tdf['id'].str.split('_').str[0]
tdf['city'] = tdf['id'].str.split('_').str[1]
tdf['city2'] = tdf['id'].str.slice(2, 4) + '시'

# .str.replace: 치환
tdf.replace('(.*)li(.*)', r'\1~\2')
tdf.state.str.replace('.Y', '뉴욕')

# tolist(), list(): 시리즈를 리스트형으로 변환
tdf[['other', 'mathod']] = pd.DataFrame(tdf['id'].str.split('_').tolist())
tdf[['other1', 'mathod1']] = pd.DataFrame(list(tdf['id'].str.split('_')))

# 천단위 콤마 없애기
tdf['pop2'] = tdf['pop'].str.replace(',', '')

# 문자형 컬럼의 숫자형 변환
tdf['pop2'].astype('int64')    # 개별지정:astype({'col_str_1':int, 'col_str_2':np.float})
tdf['pop2'] = pd.to_numeric(tdf['pop2'])
tdf['pop3'] = tdf['pop2'] / 1234

# 천단위 콤마 추가
tdf['pop2'].apply(lambda x: format(x, ','))

# 소수점 자리수 변경
tdf['pop3'].round(3)
tdf['pop3'].apply(lambda x: '{:.2f}'.format(x))

# 순위
tdf['rank'] = tdf['pop2'].rank(method='min', ascending=False)
tdf['rank2'] = tdf['pop2'].astype('int64').rank(method='min', ascending=False)
tdf['c_rank'] = tdf.groupby('class')['pop2'].rank(method='min', ascending=False)

# #  시계열 자료 다루기 ----------

# # DatetimeIndex 인덱스: timestamp 형식의 인덱스 생성 ==========

# 판다스에서 시계열 자료를 편리하게 다루려면 인덱스를 DatetimeIndex 자료형으로 만들어야 함

# DatetimeIndex 인덱스 생성
# 1. pd.to_datetime 함수:  날짜/시간 문자열을 datetime 자료형으로 바꾼 후 DatetimeIndex 자료형 생성
# 2. pd.date_range 함수: 시작일과 종료일 또는 시작일과 기간으로 범위 내의 인덱스를 생성
# 이렇게 만들어진 인덱스를 사용하여 시리즈나 데이터프레임을 생성하면 된다.

# # 많이 사용되는 freq 인수값
# s: 초
# T: 분
# H: 시간
# D: 일
# B: 주말이 아닌 평일
# W: 각 주의 일요일
# W-MON: 각 주의 월요일
# M: 각 달의 마지막 날
# MS: 각 달의 첫날
# BM: 주말이 아닌 평일 중에서 각 달의 마지막 날
# BMS: 주말이 아닌 평일 중에서 각 달의 첫날
# WOM-2THU: 각 달의 두번째 목요일
# Q-JAN: 각 분기의 첫달의 마지막 날
# Q-DEC: 각 분기의 마지막 달의 마지막 날

import pandas as pd
import numpy as np

date_str = ["2018, 1, 1", "2018, 1, 4", "2018, 1, 5", "2018, 1, 6"]
idx = pd.to_datetime(date_str)

pd.date_range("2018-4-1", "2018-4-30", freq="W-MON")
pd.date_range(start="2018-4-1", periods=30)

np.random.seed(0)
s = pd.Series(np.random.randn(4), index=idx)
s

# (date + time)컬럼 -> 인덱스
Date = ['01/12/2021', '30/1/2021', '15/3/2022', '30/08/2024', '30/09/2023']
Time = ['09.01.00', '12.01.01', '15.01.02', '23.10.1', '12.30.00']
val = [3500, 5000, 2300, 5600, 4000]
df = pd.DataFrame({'Date': Date, 'Time': Time, 'val': val})
df.index = pd.to_datetime(df.Date + ' ' + df.Time, format='%d/%m/%Y %H.%M.%S')
print(df)

# df.set_index('기존dt열', drop=False, inplace=True)

# # resample 연산: 시간 간격을 재조정 ==========

# 업-샘플링(시간 구간이 작아져서 데이터 양이 증가)의 경우에는 새로운 인덱스에 Nan 이 추가되므로
#          실제로 존재하지 않는 데이터를 만들어야 함(ffill, bfill 메서드를 이용)
# 다운-샘플링(시간 구간이 커져서 데이터 양이 감소)의 경우에는 원래의 데이터값들이 묶이기 때문에 대표값을 구해야 함
# ohlc 메서드는 구간의 시고저종(open, high, low, close)값을 구함

ts = pd.DataFrame(np.random.randn(100), columns=['value'], index=pd.date_range(
                  "2018-1-1", periods=100, freq="D"))
ts.tail(20)

ts.resample('10H').ffill()
ts.resample('M').first()
ts.resample('W').ohlc()

# # dt 접근자: datetime형 시리즈의 몇가지 속성과 메서드 사용 가능 ==========

# 년, 월, 일, 요일 정보(year, month, day, weekday) 추출
df = pd.DataFrame({'Birth': ['2019-01-01 09:10:00', '2019-01-08 09:20:30', '2019-02-01 10:20:00',
                             '2019-02-02 11:40:50', '2019-02-28 15:10:20', '2019-04-10 19:20:50',
                             '2019-06-30 21:20:50', '2019-07-20 23:30:59']})

df['Birth'] = pd.to_datetime(df['Birth'], format='%Y-%m-%d %H:%M:%S', errors='raise')

df['Birth_date'] = df['Birth'].dt.date                    # YYYY-MM-DD(문자)
df['Birth_year'] = df['Birth'].dt.year                    # 연(4자리숫자)
df['Birth_month'] = df['Birth'].dt.month                  # 월(숫자)
df['Birth_month_name'] = df['Birth'].dt.month_name()      # 월(문자)
df['Birth_day'] = df['Birth'].dt.day                      # 일(숫자)
df['Birth_time'] = df['Birth'].dt.time                    # HH:MM:SS(문자)
df['Birth_hour'] = df['Birth'].dt.hour                    # 시(숫자)
df['Birth_minute'] = df['Birth'].dt.minute                # 분(숫자)
df['Birth_second'] = df['Birth'].dt.second                # 초(숫자)

df['Birth_quarter'] = df['Birth'].dt.quarter              # 분기(숫자)
df['Birth_weekday_name'] = df['Birth'].dt.day_name()      # 요일이름(문자)
df['Birth_weekday'] = df['Birth'].dt.weekday              # 요일숫자(0-월, 1-화) (=dayofweek)
df['Birth_weekofyear'] = df['Birth'].dt.weekofyear        # 연 기준 몇주째(숫자) (=week)
df['Birth_dayofyear'] = df['Birth'].dt.dayofyear          # 연 기준 몇일째(숫자)
df['Birth_days_in_month'] = df['Birth'].dt.days_in_month  # 월 일수(숫자) (=daysinmonth)

df['Birth_is_leap_year'] = df['Birth'].dt.is_leap_year          # 윤년 여부
df['Birth_is_month_start'] = df['Birth'].dt.is_month_start      # 월 시작일 여부
df['Birth_is_month_end'] = df['Birth'].dt.is_month_end          # 월 마지막일 여부
df['Birth_is_quarter_start'] = df['Birth'].dt.is_quarter_start  # 분기 시작일 여부
df['Birth_is_quarter_end'] = df['Birth'].dt.is_quarter_end      # 분기 마지막일 여부
df['Birth_is_year_start'] = df['Birth'].dt.is_year_start        # 연 시작일 여부
df['Birth_is_year_end'] = df['Birth'].dt.is_year_end            # 연 마지막일 여부
print(df)

# # 시계열 원소 다루기 ==========

# 현재 일시 및 요소 추출
pd_today = pd.Timestamp.today()
pd_ts = pd.Timestamp.now()
pd_ts.date()
pd_ts.time()
pd_ts.year    # .month.day.hour.minute.second.microsecond

# 인덱스로 데이터 행 추출
# 판다스가 제공하는 고급 인덱서(loc, iloc)
df.iloc[:2]
df.loc['2024-08-30']
df.loc['2020':'2022']
df.loc['2021-02':'2023-02']

df['2021']
df['2022-03']
df['2022-02-25':'2023-12-31']

df[(df.index.year == 2022) | (df.index.year == 2023)]
df[df.index <= '2022-12-01']
df[df.index >= pd.datetime(2022, 12, 1)]

# 인덱스 중복(개수) 확인
df.groupby(level=0).count()

# 일시 사칙연산
pd_ts + pd.Timedelta(days=1, hours=1, minutes=15, weeks=1)
pd_ts + 5 * pd.Timedelta(days=1)                        # 5days
pd_ts - pd.Timestamp(2019, 5, 2, 13, 30, 50, 100000)    # 날짜, 초, 마이크로초 차이 정보 추출

# 인덱스별 데이터 집계
df.groupby(level=0).agg(['size', 'sum', 'mean', 'min', 'max'])

# 등간격 구간별 데이터 집계
df2 = pd.DataFrame()
df2['val_6M_sum'] = df.val.resample('6M').sum()
df2['val_6M_cumsum'] = df.val.resample('6M').sum().cumsum()
df2['val_6M_max'] = df.val.resample('6M').max()
df2['val_6M_min'] = df.val.resample('6M').min()
df2['val_6M_mean'] = df.val.resample('6M').mean()
df2['val_6M_median'] = df.val.resample('6M').median()
df2['val_6M_first'] = df.val.resample('6M').first()
df2['val_6M_last'] = df.val.resample('6M').last()
df2['val_6M_var'] = df.val.resample('6M').var().fillna(0)
# df2['val_6M_stddev'] = np.sqrt(df.val.resample('6M').var())  # 표준편차
print(df2)

# 선형적 기간 집계
ts['value'].resample('M').count()

# 순환적 기간 집계 (시간, 요일 등 반복개념)
ts.index.day_name().value_counts()
ts.index.day.value_counts().sort_index()

# 일시 -> 문자열
str(pd_ts)[:10]                            # str()
pd_ts.strftime('%y년 %m월 %d일 %I시 %M분')  # .strftime(): 포멧지정

# 문자열 -> 일시
pd.to_datetime(['Dec 22, 2019 01:30:59 PM', '20100505', '200412'])  # 자동변환
pd.to_datetime('200412', format='%y%m%d')                           # 특수양식(yymmdd)
pd.to_datetime('2019년 4월 10일', format='%Y년 %m월 %d일')           # 특수양식(한글)
# %Y(2020), %y(20), %m(01~12), %d(01~31)
# %H(01~23), %I(01~12), %M(01~59), %S(00~61)
# %w(0일~6토), %U(일요일기준누적주), %W(월요일기준누적주)
# %B(Jauary), %b(Jan), %A(Sunday), %a(Sun)

# 숫자 -> 일시
print(pd.Timestamp(2019, 5, 2, 13, 30, 50, 100000))

# Timestamp -> datetime 변환
# Timestamp(1970~2038;초단위), datetime(1000~9999;고정,시간대)
print(pd.datetime(2019, 5, 2, 13, 30, 50, 100000))
pd.Timestamp('2019-5-2,13:30:50').to_pydatetime()

# 시간대 설정 및 변경
ts_seoul = pd_ts.tz_localize('Asia/Seoul')   # 시간대 설정
ts_UTC = ts_seoul.tz_convert('UTC')

"""
시계열 데이터 분석
"""

import pandas as pd
import datetime
import re


# # 데이터 로딩 ----------
df = pd.read_excel('c:/Users/Administrator/Desktop/2022.xlsx')
df.shape      # 총데이터수 18,907개

# 결측치 확인
missing_df = df.isnull()
for col in missing_df.columns:
    missing_count = missing_df[col].value_counts()
    try:
        print(col, ': ', missing_count[True])
    except BaseException:
        print(col, ': ', 0)

# # 데이터 정제

# 특정 기록자 제외
ndf = df[~df.확인자.str.contains('점검조|테스트', na=False)]
ndf = ndf[~ndf.처리자.str.contains('점검조|테스트', na=False)]
ndf.shape     # 잔여데이터 17,887개

# 통신단절 제외
ndf = ndf[~ndf.경보명.str.contains('통신단절|Disconnect|Timeout|원격관제모듈', na=False)]
ndf.shape     # 잔여데이터 16,991개

# 리테일 제외
ndf = ndf[~df.경보명.str.contains('GS수퍼|EMS')]
ndf.shape     # 잔여데이터 2,384개

# 동일자 발생알람 제외
ndf['n발생일자'] = ndf['발생일시'].str[0:10]
ndf = ndf.drop_duplicates(subset=['경보명', 'n발생일자'])
ndf.shape     # 잔여데이터 1,184개

ndf.head(2)

# k = ndf.경보명.str.extract(r'(.+)_(.+)_(.+)').get(1).value_counts().index.unique()

tdf_fire = ndf[ndf.경보명.str.contains('화재', na=False)].경보명
tdf_not_fire = ndf[~ndf.경보명.str.contains('화재', na=False)].경보명
tdf_water = tdf_not_fire[tdf_not_fire.str.contains('고수위|저수위|펌프|LOW|정화조|누수감지기', na=False)]
tdf_not_water = tdf_not_fire[~tdf_not_fire.str.contains('고수위|저수위|펌프|LOW|정화조|누수감지기', na=False)]
tdf_elec = tdf_not_water[tdf_not_water.str.contains('발전기|누전경보기|ALT|ACB|CTTS', na=False)]
tdf_not_elec = tdf_not_water[~tdf_not_water.str.contains('발전기|누전경보기|ALT|ACB|CTTS', na=False)]

tdf_fire.shape
tdf_water.shape
tdf_elec.shape
tdf_not_elec.shape


# # 데이터 샘플링 ----------

ndf2 = ndf[['발생일시', '경보명', '처리내용']]
ndf2.head(2)
ndf2.dtypes
ndf2['발생일시'] = pd.to_datetime(ndf2['발생일시'])
ndf2.set_index('발생일시', drop=False, inplace=True)
ndf2.shape

# 선형적 기간 집계
day_count = ndf2['경보명'].resample('D').count()
month_count = ndf2['경보명'].resample('M').count()

# 순환적 집계
week_count = ndf2.index.weekday.value_counts().sort_index()
hour_count = ndf2.index.hour.value_counts().sort_index()

# 시각화
from matplotlib import font_manager, rc
font_name = font_manager.FontProperties(fname="D:/2022/_Py/data/malgun.ttf").get_name()
rc('font', family=font_name)

day_count.plot(kind='area')
month_count.plot(kind='bar')
week_count.plot(kind='bar', xlabel='Mon(0) ~ Sun(6)')
hour_count.plot(kind='bar')
import calplot
calplot.calplot(day_count, cmap='YlGn')

import seaborn as sns
ndf2['week'] = ndf2.index.day_name()
ndf2['hour'] = ndf2.index.hour
ndf_week_hour = ndf2.groupby(['week', 'hour']).size()
ndf_table = ndf_week_hour.rename_axis(['Weekday', 'Hour']).unstack('Weekday')
days=['Monday', 'Tuesday', 'Wednesday', 'Tuesday', 'Friday', 'Saturday', 'Sunday']
ndf_table_sort = ndf_table.reindex(columns=days).sort_index(ascending=False)
sns.heatmap(ndf_table_sort, cmap='YlGn')

# 일별 시계열 분해
from statsmodels.tsa.seasonal import seasonal_decompose
ts = ndf2['경보명'].resample('D').count()
result = seasonal_decompose(ts, model='Additive')
result.plot()
plt.show()

#######################################
# 데이터프레임을 합치거나 구조 변경하는 법
#######################################

import pandas as pd
df_1 = pd.DataFrame({'년도': ['2021', '2021', None, '2022'],
                     '지역': ['서울', '경기', '부산', '서울'],
                     '인구': [59000, 72000, 40500, 25000],
                     '세대수': [2500, 2000, 1900, None],
                     '기온': [19.5, 20.6, None, 25.5]},
                    index=[i for i in range(0, 8, 2)])

df_2 = pd.DataFrame([['2021', '서울', 59000, 19.5],
                     ['2022', '서울', 25000, 25.5],
                     ['2021', '경기', 72000, 20.6],
                     ['2021', '광주', None, 27.1],
                     ['2022', '대구', 35000, 29.1],
                     ['2022', '경기', 88000, 23.1],
                     ['2021', '제주', 39000, 36.5]],
                    columns=['년도', '지역', '인구', '기온'])

print(df_1, df_2, sep='\n\n')

# 이어 붙이기
pd.concat([df_1, df_2])                             # default: join='outner'(합집합), axis=0
pd.concat([df_1, df_2], ignore_index=True)          # ignore_index = True(새로운 정수형 위치 인덱스 설정)
pd.concat([df_1, df_2], join='inner', axis=1)       # 'inner'(교집합)은 행|열이름 기준
pd.concat([df_2, df_1['세대수']], join='inner', axis=1)  # 데이터프레임과 시리즈 이어붙이기

df_1.append(df_2, ignore_index=True)
df_1.join(df_2, how='left', lsuffix='_left', rsuffix='_right')  # 행인덱스를 기준으로 붙이기

# 특정 열 기준에 의한 병합
pd.merge(df_1, df_2)                         # default: how='inner', on=None(공통되는모든열을merge-key)
pd.merge(df_1, df_2, how='left', on='지역')
pd.merge(df_1, df_2, how='outer', on='지역', indicator='source')
# how='inner'                            # inner, outer, left, rigth
# on=None, left_on=None, right_on=None   # merge-key: None(공통되는 모든 열), 공통되는특정열명칭, 데이터프레임별기준열명칭
# left_index=False, right_index=False    # merge-key로 index 사용
# suffixes=('_좌', '_우')                # 중복 변수명에 붙이는 접미사 설정

# 덮어쓰기 (None 제외)
# df1.update(df2, join='left'(왼쪽옵션만존재), overwrite=True, filter_func=None, errors='ignore')
# join 옵션은 left만 존재하고, 오른쪽 병합df2가 None이면 기존 df1 데이터를 유지
df_2.update(df_1)
print(df_2)

# 데이터프레임을 변수명과 관측치로 길게(rows) 풀어 쓰기
pd.melt(df_1)
pd.melt(df_1, id_vars=['년도', '지역'], var_name='변수명', value_name='관측치')

# 복제
df3 = df2[:]
df4 = df2[:]
df5 = df2[:]


#######################################
# 데이터프레임에 함수를 매핑하는 법
#######################################

# Numpy
배열(array)을 사용하면 적은 메모리로 많은 데이터를 빠르게 처리할 수 있지만 파이썬은 자체적으로 배열 자료형을 제공하지 않는다.
따라서 배열을 구현할 수 있는 Numpy를 임포트해야 한다.
배열은 리스트와 비슷하지만 모든 원소가 같은 자료형이어야 하며 원소의 갯수를 바꿀 수 없는 특징이 있다.

벡터화 연산(vectorized operation)을 지원
data = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
import numpy as np
np.array(data) * 2

# 판다스 시리즈객체
시리즈 객체는 라벨 값에 의해 인덱싱이 가능하므로 실질적으로 인덱스 라벨 값을 키(key)로 가지는 딕셔너리 자료형과 같다고 볼 수 있다

넘파이 배열처럼 판다스의 시리즈도 벡터화 연산을 할 수 있다.

# 카테고리 세기
시리즈의 값이 정수, 문자열, 카테고리 값인 경우에는 value_counts 메서드로 각각의 값이 나온 횟수를 셀 수 있다. 그러나, 데이터프레임에는 value_counts 메서드가 없으므로 각 열마다 별도로 적용해야 한다.

# 정렬
데이터를 정렬하려면 sort_index 메서드 sort_values 메서드를 사용한다. sort_index 메서드는 인덱스 값을 기준으로, sort_values 메서드는 데이터 값을 기준으로 정렬한다.
s.value_counts().sort_index()
s.sort_values(ascending=False)
데이터프레임에서 sort_values 메서드를 사용하려면 by 인수로 정렬 기준이 되는 열을 지정해 주어야 한다.
df.sort_values(by=['기준1', '기준2'])

# 합계
행과 열의 합계를 구할 때는 sum(axis) 메서드를 사용
df2.sum(axis=1)  # 행방향
df2["RowSum"] = df2.sum(axis=1)  # 오른쪽 끝에 집계열 추가
df2.sum()  # 열방향
df2.loc["ColTotal", :] = df2.sum()  # 맨밑에 집계행 추가

# apply 변환
행이나 열 단위로 더 복잡한 처리를 하고 싶을 때는 apply 메서드를 사용
df3.apply(lambda x: x.max() - x.min(), axis=1)

# 실수 값을 카테고리 값으로 변환
실수 값을 크기 기준으로 하여 카테고리 값으로 변환하여 Categorical 클래스 객체로 반환
cut: 경계값을 기준으로 구간 나눔
qcut: 동일 갯수를 기준으로 구간 나눔

ages = [0, 1, 19, 20, 21, 69, 70, 71, 101]
bins = [1, 20, 30, 50, 70, 100]
labels = ["미성년자", "청년", "중년", "장년", "노년"]
cats = pd.cut(ages, bins, labels=labels)
cats
bins 은 앞쪽경계값은 미포함, 뒷쪽경계값은 포함, 영역을 넘는 값은 NaN으로 처리
labels 은 구간명
반환값은 문자열이 아니므로 문자열 사용을 위해서는 astype 메서드 사용 필요
df4.age_cat.astype(str)

data = np.random.randn(1000)
cats = pd.qcut(data, 4, labels=["Q1", "Q2", "Q3", "Q4"])  # 4개 구간을 생성
cats


# 산술연산--------------

# --------------
import pandas as pd
df = pd.DataFrame({'id': [1, 2, 10, 20, 100, 200],
                   'score': [3002, 2903, 1004, 3920, 1090, 1923],
                   'name': ['aaa', 'bbb', 'ccc', 'ddd', 'eee', 'fff']})


def add_10(n):
    return n + 10


# lambda input1, input2, ... : 표현식
r1 = lambda x: '%05d' % x
r2 = lambda x: "{:0>6d}".format(x)
r3 = lambda x: x.max() - x.min()
r4 = lambda x: pd.Series([x.min(), x.max()], index=['min', 'max'])
r5 = lambda x: ':'.join(x)
r6 = lambda x, y: x * y

# 시리즈 '원소'에 함수 매핑: apply()
df['id_2'] = df['id'].apply(add_10)
df['id_3'] = df['id'].apply(lambda x: add_10(x))
df['id'].apply(r1)
df['id'].apply(r2)

# 데이터프레임 '원소'에 함수 매핑: applymap()
df2 = df.loc[:, ['id', 'score']]
df2 = df2.applymap(add_10)
df2 = df2.applymap(r1)

# 데이터프레임 각 '열|행'에 함수 매핑: 함수에 따라 반환객체 다름
df[['id', 'score']].apply(r3)          # 열에 매핑 axis=0
df[['id', 'score']].apply(r3, axis=1)  # 행에 매핑 axis=1
df[['id', 'score']].apply(r4)
df['id_name'] = df[['id_2', 'name']].astype(str).apply(r5, axis=1)

# '데이터프레임' 객체에 함수 매핑: pipe()
df.pipe(lambda x: x.isnull())
df.pipe(lambda x: x.sum())
df.pipe(lambda x: x.count().sum())

# map(func, 인자) 함수 : 리스트 또는 시리즈 반환
df['id'].map(r2)
list(df['id'].map(r2))
list(map(r6, (100, 50), (200, 100)))


#######################################
# 그룹객체 다루기
#######################################

# 1. 피벗테이블 --------------

# 피봇테이블은 두 개 이상의 열을 행인덱스와 열인덱스로 하여 특정 데이터열의 데이터를 펼쳐놓은 것이다
# 행인덱스와 열인덱스의 조건을 만족하는 데이터가 2개 이상인 경우는 에러가 발생한다

import pandas as pd
data = {
    "도시": ["서울", "서울", "서울", "부산", "부산", "부산", "인천", "인천"],
    "연도": ["2015", "2010", "2005", "2015", "2010", "2005", "2015", "2010"],
    "인구": [9904312, 9631482, 9762546, 3448737, 3393191, 3512547, 2890451, 263203],
    "지역": ["수도권", "수도권", "수도권", "경상권", "경상권", "경상권", "수도권", "수도권"]
}
columns = ["도시", "연도", "인구", "지역"]
df1 = pd.DataFrame(data, columns=columns)
print(df1)
df1.pivot("도시", "연도", "인구")
# 피봇테이블은 set_index 과 unstack 을 사용해서 만들 수도 있음
df1.set_index(["도시", "연도"])[["인구"]].unstack()

df_1 = pd.DataFrame({'년도': ['2021', '2021', None, '2022'],
                     '지역': ['서울', '경기', '부산', '서울'],
                     '인구': [59000, 72000, 40500, 25000],
                     '세대수': [2500, 2000, 1900, None],
                     '기온': [19.5, 20.6, None, 25.5]},
                    index=[i for i in range(0, 8, 2)])

# 피벗테이블은 피라미드의 밑에서부터 c->v->a 순으로 분리해 감
df_pivot = pd.pivot_table(df_1,                    # 데이터프레임
                          index=['년도', '세대수'],                 # 행 위치에 들어갈 열
                          columns='지역',                          # 열 위치에 들어갈 열
                          values=['인구', '기온'],                  # 데이터로 사용할 열
                          aggfunc=['sum', 'mean'],
                          # aggfunc={'인구': 'sum', '기온': 'mean'},  # 데이터 집계 함수 (디폴트는 'mean')
                          fill_value=0,                            # NaN 채우기
                          margins=True,                            # 행별,열별 합계
                          margins_name='합계')

# 행 인덱스 접근: xs 인덱서
df_pivot.xs('2021')
df_pivot.xs(2500, level=1)                 # 레벨을 직접 지정
df_pivot.xs(2500, level='세대수')
df_pivot.xs(('2021', 2500))                # 멀티레벨은 튜플로 인수 전달, 시리즈형 반환
df_pivot.xs(('2021', 2500), level=(0, 1))  # 데이터프레임형 반환

# 열 인덱스 접근: axis=1
df_pivot.xs('기온', level=0, axis=1)
df_pivot.xs(('기온', '서울'), level=(0, 1), axis=1)
df_pivot['기온']
df_pivot['기온', '서울']

# 사용자 정렬 : 레벨(index, column)
df_pivot.reorder_levels(['세대수', '년도'])
df_pivot.reorder_levels([1, 0], axis=1)
df_pivot.swaplevel(0, 1)
df_pivot.swaplevel(0, 1, axis=1)

# 컬럼과 인덱스간 전환(그룹연산용)
df_pivot.stack()
df_pivot.stack(level=0).unstack()

# 사용자 정렬 : 인덱스(그래프에 영향 없음)
year_order = ['합계', '2022', '2021']
df_pivot.reindex(year_order, level=0, axis=0)

# 사용자 정렬 : 인덱스2(sns 그래프용)
df2 = df_pivot.reset_index().set_index('년도')
df2.index = pd.CategoricalIndex(df2.index, categories=year_order)
df2.sort_index(level=0, inplace=False)

# 데이터의 갯수를 세는 피벗테이블
df3 = (pd.DataFrame({'X': ['X1', 'X1', 'X1', 'X1'],
                     'Y': ['Y2', 'Y1', 'Y1', 'Y1'],
                     'Z': ['Z3', 'Z1', 'Z1', 'Z2']}))
print(df3)
# 카운트
pd.pivot_table(df3, values='X', index='Y', columns='Z', aggfunc='count')
# 유니크카운트
df3.pivot_table(values='X', index='Y', columns='Z', aggfunc=lambda x: len(x.unique()))
df3.pivot_table(values='X', index='Y', columns='Z', aggfunc=pd.Series.nunique)

# 인덱스 설정 및 제거
set_index : 기존의 행 인덱스를 제거하고 데이터 열 중 하나를 인덱스로 설정
reset_index : 기존의 행 인덱스를 제거하고 인덱스를 데이터 열로 추가

# 다중 인덱스
index 또는 columns 을 리스트의 리스트(행렬) 형태로 인덱스 설정
df = pd.DataFrame(np.round(np.random.randn(6, 4), 2),
                   columns=[["A", "A", "B", "B"],
                            ["C", "D", "C", "D"]],
                   index=[["M", "M", "M", "F", "F", "F"],
                          ["id_" + str(i + 1) for i in range(3)] * 2])
df.columns.names = ["Cidx1", "Cidx2"]
df.index.names = ["Ridx1", "Ridx2"]
df

# 행 인덱스와 열 인덱스 교환
stack    열 인덱스 -> 행 인덱스로 변환
unstack 행 인덱스 -> 열 인덱스로 변환
df.stack("Cidx1")
df.stack(1)
df.unstack("Ridx2")
df.unstack(0)

# 다중 인덱스의 인덱싱(튜플처리)
df[("B", "C")]
df.loc[("M", "id_2"), ("A", "D")]
df.iloc[2, 3]  # iloc 인덱서는 튜플 형태의 다중인덱스를 사용할 수 없다
df["A"]  # 하나의 레벨 값만 넣으면 다중 인덱스 중 가장 상위값을 반환
df.loc[("All", "All"), :] = df.sum()
df.loc[(slice(None), "id_1"), :]  # 특정 레벨의 모든 인덱스 값을 인덱싱할 때는 슬라이스를 사용한다. 다만 다중 인덱스의 튜플 내에서는 : 슬라이스 기호를 사용할 수 없으므로 slice(None) 값을 사용한다.

# 다중 인덱스의 인덱스 레벨 맞교환
df.swaplevel("Ridx1", "Ridx2")
df.swaplevel("Cidx1", "Cidx2", 1)  # 0일 때 행 인덱스, 1일 때 열 인덱스

# 다중 인덱스의 정렬
df.sort_index(level=0)
df.sort_index(axis=1, level=0)


# # 2. groupby클래스객체 ----------

# # 지정 조건에 맞는 데이터가 하나 이상이면 그룹분석을 통해 해당 그룹의 대표값 구해야 함
# size, count: 그룹 데이터의 갯수
# mean, median, min, max: 그룹 데이터의 평균, 중앙값, 최소, 최대
# sum, prod, std, var, quantile : 그룹 데이터의 합계, 곱, 표준편차, 분산, 사분위수
# first, last: 그룹 데이터 중 가장 첫번째 데이터와 가장 나중 데이터
# agg, aggregate: 사용자 함수를 만들어서 전달
# describe: 기본 통계값들을 데이터프레임으로 출력
# apply: 사용자 함수를 사용하여 데이터프레임으로 출력
# transform: 그룹의 대표값을 만드는 것이 아니라 그룹별 계산을 통하여 데이터 자체를 변형

df = pd.DataFrame({
    'key1': ['A', 'A', 'B', 'B', 'A'],
    'key2': ['one', 'two', 'one', 'two', 'one'],
    'data1': [1, 2, 3, 4, 5],
    'data2': [10, 20, 30, 40, 50]
})

# # key1의 값(A 또는 B)에 따른 data1의 합계 구하기
# 순차법
groups = df.groupby(df.key1)
groups.groups  # 그룹데이터의 인덱스 확인
groups.sum()
# 직접법(1): data1을 키값으로 그룹핑 후 합산 적용
df.data1.groupby(df.key1).sum()
# 직접법(2): 키값으로 df를 그룹핑 후 data1만을 선택하여 합산 적용
df.groupby(df.key1)["data1"].sum()
# 직접법(3): 키값으로 df를 그룹핑 후 합산 후 data1을 선택
df.groupby(df.key1).sum()["data1"]

# 복합 키 (key1, key2) 는 리스트를 사용한다.
df2.data1.groupby([df2.key1, df2.key2]).sum()
df2.data1.groupby([df2["key1"], df2["key2"]]).sum().unstack("key2")  # 피벗 형태 변환

agg 메서드는 필요한 그룹연산 메서드를 만들어 사용한다

describe 메서드는 다양한 기술 통계(descriptive statistics)값을 한 번에 구하므로 그룹별로 하나의 스칼라 값이 아니라 하나의 데이터프레임이 생성된다

apply 메서드는 describe 메서드처럼 하나의 그룹에 대하여 하나의 대표값(스칼라 값)을 구하는 게 아니라 데이터프레임을 만들 수 있다

transform 메서드는 그룹별 대표값을 만드는 것이 아니라 그룹별 계산을 통해 데이터프레임 자체를 변화시킨다

# # pivot_table
Pandas는 pivot 명령과 groupby 명령의 중간 성격을 가지는 pivot_table 명령도 제공한다. pivot_table은 groupby 명령처럼 그룹분석을 하지만 최종적으로는 pivot 명령처럼 피봇테이블을 만든다.

pivot_table(data, values=None, index=None, columns=None, aggfunc='mean', fill_value=None, margins=False, margins_name='All')
data: 분석할 데이터프레임 (메서드일 때는 필요하지 않음)
values: 분석할 데이터프레임에서 분석할 열
index: 행 인덱스로 들어갈 키 열 또는 키 열의 리스트
columns: 열 인덱스로 들어갈 키 열 또는 키 열의 리스트
aggfunc: 분석 메서드
fill_value: NaN 대체 값
margins: 모든 데이터를 분석한 결과를 오른쪽과 아래에 붙일지 여부
margins_name: 마진 열(행)의 이름

pivot_table를 메서드로 사용할 때는 객체 자체가 데이터가 되므로 data 인수가 필요하지 않다.
df1.pivot_table("인구", "도시", "연도")

import pandas as pd
import seaborn as sns
titanic = sns.load_dataset('titanic')
df = titanic.loc[:, ['age', 'sex', 'class', 'fare', 'survived']]

# 한개 열을 기준으로 그룹화
grouped = df.groupby('class')   # class열은 3개('first', ' second', 'third')의 튜플형태 그룹이 존재
for i in grouped:
    print(type(i))

# 그룹객체는 그룹이름(key)과 그룹으로 쪼개진 데이터(group)를 튜플형태로 묶어서 가지고 있다. 또한 기존의 행 인덱스번호를 그대로 가진다.
for key, group in grouped:
    print('* key : ', key)
    print('* number : ', len(group))
    print(group.head())
    print('\n')

# # 2. 그룹객체에 대한 일반 연산 ----------

# 2-1. 그룹별 평균 : 그룹별(위에 짤려진 데이터별)로 연산을 수행한다
grouped.mean()

# 2-2. 그룹 추출 : 나눠진 그룹 중 특정 그룹만 가져옴
grouped.get_group('Third')

# 2-3. 여러 열을 기준으로 그룹객체 생성
grouped_two = df.groupby(['class', 'sex'])  # [열리스트]로 전달 class그룹수(3)*sex그룹수 (2)만큼의 나눠어진 그룹(6)생성
for key, group in grouped_two:
    print('* key : ', key)
    print('* number : ', len(group))
    print(group.head())
    print('\n')

# 2-4. 멀티인데스로 나눠진 특정 그룹객체 가져오기(튜플로 전달)
grouped_two.get_group(('Third', 'female'))

# 2-5. 그룹별 엑셀파일 저장
for key, group in grouped_two:
    group.to_excel(f'{key}.xlsx')

# # 3. 그룹연산 메서드 --------------

# 3-1. 판다스 내장함수: 연산이 가능한 숫자형 행열에 NaN 제외하고 선택적 연산 수행
grouped = df.groupby('class')    # 그룹객체 생성 후 연산 진행
grouped.sum()                    # mean(),max(),min(),sum(),count()
grouped.quantile([0.25, 0.75])   # size(),var(),std(),median() 등
grouped.rank(method='min', ascending=False)

grouped.fare.sum()              # 각 그룹의 특정 열에 대한 집계

# 3-2. agg(): 여러열에 여러함수를 적용가능


# 사용자 정의함수
def min_max(x):
    return x.max() - x.min()


grouped.agg(min_max)

# 모든열에 하나의 함수를 매핑 : group객체.agg(함수)
grouped.agg('mean')

# 모든열에 여러 함수를 매핑 : group객체.agg([함수1,함수2,함수3,…])
grouped.agg(['count', 'mean', 'sum'])

# 각 열마다 다른 함수를 매핑 : group객체.agg({‘열1’: 함수1, ‘열2’:함수2, …})
grouped.agg({'age': ['size', 'mean', 'std'], 'fare': min_max})

# 3-3. transform(): 그룹별로 매핑함수를 적용한 후, 원래 데이터프레임의 인덱스순서로 결과값 반환
grouped.transform('mean')
grouped.transform(min_max)

# 3-4. filter(): 그룹객체를 필터링
grouped.filter(lambda x: len(x) >= 200)       # 그룹별 row수가 200개 이상인 그룹만 선택
grouped.filter(lambda x: x.age.mean() < 30)

# 3-5. apply(): 그룹객체에 대한 함수 매핑
grouped.apply('describe')
grouped.apply(lambda x: x.describe())
grouped.apply(lambda x: x.describe()).reset_index()   # 계층 인덱스를 컬럼으로 흡수

# 연산 결과값 추출
a_gup = grouped_two.apply('mean')
a_gup[2:4]
a_gup.loc['Second']
a_gup.loc[('Second', 'male')]   # 튜플로 인수 전달
a_gup.loc['Second', 'male']
a_gup.xs('male', level='sex')   # 멀티인덱스 2단계 기준에서 추출

# # 4. 멀티인덱서: .xs --------------

# 기본 데이터프레임의 인덱싱 함수(loc, iloc)를 이용하려면 큰 그룹부터 순차적으로 인덱싱을 해야하지만,
# 멀티인덱서(.xs)를 이용하면 수준(level)에 따른 인덱싱이 가능

# 4-1. 행의 멀티인덱스를 인덱싱
import pandas as pd
import seaborn as sns

df = sns.load_dataset('titanic')[['age', 'sex', 'class', 'fare', 'survived']]
grouped = df.groupby(['class', 'sex'])   # 멀티인덱스
gdf = grouped.mean()

print(gdf)
print(gdf.loc['First'])

print(gdf.xs('male', level='sex'))                   # 멀티인덱서 sex가 두번째 그룹범주이므로 ‘sex’ 대신 1을 입력해도 됨
gdf.xs(('First', 'female'))                          # flpat64객체반환
gdf.xs(('First', 'female'), level=['class', 'sex'])  # level을 지정해주면, 데이터프레임객체반환

# 4-2. 열의 멀티인덱스를 인덱싱: 행인덱스 인덱싱방법과 동일하며 axis=1옵션만 추가
gdf.xs('mean', axis=1)
gdf.xs(('mean', 'age'), axis=1)

# 4-3. 멀티인덱스 해제
# 행인 경우는 기존 인덱스는 열로 이동 시키고, 정수형 위치 인덱스로 초기화
print(gdf.reset_index())

# 컬럼이 멀티인덱스인 경우는 컬럼명을 새로 지정해주면 멀티인덱스가 해제된다.
gdf2 = df.groupby('class').agg(['mean', 'max'])[['age', 'fare']]
print(gdf2)
gdf2.columns = ['age_mean', 'age_max', 'fare_mean', 'fare_max']
print(gdf2)


# 6. 판다스 화면 설정 --------------

# IPython 디스플레이 설정
pd.get_option()    # 옵션 확인
pd.set_option()    # 옵션 설정
pd.reset_option()  # 옵션 초기화

pd.set_option('display.max_rows', None)                  # 출력할 행의 최대 개수
pd.set_option('display.max_columns', 10)                 # 출력할 열의 최대 개수
pd.set_option('display.max_colwidth', 20)                # 개별 열의 너비
pd.set_option('display.unicode.east_asian_width', True)  # 유니코드 사용 너비 조정


# ------------------------------------
# PYINSTALLER commands
# ------------------------------------

# dist폴더에 exe 파일 생성
# cmd > pyinstaller -F 파일명.py  #'하나의 파일'지정 옵션
# cmd > pyinstaller -w 파일명.py  #gui등 터미널없이 윈도우모드로 실행 옵션
# cmd > pyinstaller -w --add-data '.\gui_basic\*.png;gui_basic'
#       --add-data '.\gui_basic\*2.png;gui_basic2' -F -i 'test.ico' 파일명.py
#       #파일에 필요한 그림파일;폴더위치'로 추가 및 실행파일 icon 삽입 명령

# [option]
# --onefile        하나의 실행파일로 생성
# --noconsole      콘솔창을 안 띄우고 실행
# --icon=test.ico  실행파일의 아이콘 변경(https://icoconvert.com/)

# ※ ERROR : No module named 'pkg_resources.py2_warn'
# pip uninstall pyinstaller
# pip install https://github.com/pyinstaller/pyinstaller/archive/develop.zip

import os


def resource_path(file_name):
    try:
        base_path = sys._MEIPASS    # pyinstaller 임시 폴더(_MEIPASS)
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, file_name)


chromedriver_path = resource_path("chromedriver.exe")
driver = webdriver.Chrome(chromedriver_path)


# ------------------------------------
# CONDA commands
# ------------------------------------

'''
# 콘다 버젼 확인
conda --version

# 파이썬 버젼 확인
python --version

# 콘다 업데이트
conda update -n base conda

# 파이썬 패키지 업데이트
conda update --all

# 설치 패키지 목록
conda list

# --------------------------
# 설치 할 패키지 검색
conda search python
conda config --show channels             # 등록된 채널들 검색
conda config --add channels conda-forge  # 채널 추가
conda install -c conda-forge 패키지명     # conda-forge
conda config --set ssl_verify False      # 콘다SSL통신오류 우회(매번)

# 특정 패키지 설치
conda install python=3.5.6

# 특정 패키지 삭제
conda uninstall python

# 특정 패키지 업데이트
conda update python

# --------------------------
# 실행 환경 목록
conda env list

# 실행 환경 활성화
conda activate <myenv>
conda deactivate

# 실행 환경 패키지 목록
conda list -n <myenv>

# 실행 환경 생성
conda create -n <myenv> python=3.9
conda create --name <myenv> --file environment.yml

# 실행 환경 삭제
conda remove -n <myenv> --all

# 실행 환경 파일 저장
conda env export > environment.yml

# 실행 환경 파일 적용(주 디렉토리에 다운 로드 후)
conda env update -f environment.yml

'''

# 쥬피터 모든 변수 삭제
# %reset > y > enter
