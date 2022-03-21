"""
업체명 정제 및 홈페이지주소 크롤링
2020.9.17
"""

import openpyxl

# 홈페이지주소 크롤링


def homepage(fmname):
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    import time

    options = Options()
    options.add_argument('--disable-extensions')
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--remote-debugging-port=9515')
    options.add_argument('--disable-setuid-sandbox')

    driver = webdriver.Chrome('D:\\2021\\_Py\\MyLib\\chromedriver.exe',
                              chrome_options=options)
    driver.get('https://www.google.com')
    time.sleep(0.5)

    element = driver.find_element_by_name('q')
    element.send_keys(fmname)
    element.submit()
    time.sleep(0.5)

    # findword = driver.find_element_by_class_name('r').text
    findaddress = driver.find_element_by_tag_name('cite').text
    driver.close()
    return findaddress


# 업체명 정제

# Open
wb = openpyxl.load_workbook(
    'D:\\2021\\_Py\\CompletionCode\\워드팩스\\홍보팩스_협회목록.xlsx')

# repeat worksheets
ws_names = wb.sheetnames

for ws_name in ws_names:
    ws = wb[ws_name]

    # Read / Write
    for m in range(2, ws.max_row+1):
        a = ws.cell(row=m, column=2).value

        a = a.strip()
        if a.find('호 ') != -1:
            a = a.split('호 ')[1]
        a = a.replace('(주)', '')
        a = a.replace('㈜', '')
        a = a.replace(' ', '')

        ws.cell(row=m, column=10, value=a)
        ws.cell(row=m, column=11, value='http://' + homepage(a))

# Save
wb.save('D:\\2021\\_Py\\홍보팩스_협회목록_siteadd.xlsx')

# END
